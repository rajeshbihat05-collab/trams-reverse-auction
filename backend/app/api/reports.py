"""
Reports API and PDF/Excel export endpoints.
"""

import io
from datetime import datetime, timezone, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.database import get_db
from app.dependencies import get_admin_user
from app.models.user import User
from app.models.auction import Auction
from app.models.bid import Bid, AuctionResult
from app.models.transporter import Transporter
from app.schemas.report import MonthlyReport

router = APIRouter(prefix="/api/reports", tags=["Reports"])


@router.get("/auction-summary")
async def auction_summary_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    query = db.query(Auction)

    if start_date:
        query = query.filter(Auction.created_at >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.filter(Auction.created_at <= datetime.fromisoformat(end_date))

    auctions = query.order_by(Auction.created_at.desc()).all()

    report = []
    for a in auctions:
        lowest_bid = db.query(func.min(Bid.amount)).filter(
            Bid.auction_id == a.id, Bid.is_latest == True
        ).scalar()

        result = db.query(AuctionResult).filter(AuctionResult.auction_id == a.id).first()
        winner = None
        if result:
            w = db.query(Transporter).filter(Transporter.id == result.winner_id).first()
            winner = w.company_name if w else None

        savings = 0
        if a.reserve_price and result:
            savings = float(a.reserve_price) - float(result.winning_amount)

        report.append({
            "auction_number": a.auction_number,
            "pickup_location": a.pickup_location,
            "destination": a.destination,
            "material_type": a.material_type,
            "vehicle_type": a.vehicle_type,
            "status": a.status,
            "total_bids": a.total_bids,
            "reserve_price": float(a.reserve_price) if a.reserve_price else None,
            "lowest_bid": float(lowest_bid) if lowest_bid else None,
            "winning_amount": float(result.winning_amount) if result else None,
            "winner": winner,
            "savings": round(savings, 2),
            "loading_date": a.loading_date.isoformat() if a.loading_date else None,
            "closing_time": a.closing_time.isoformat() if a.closing_time else None,
            "created_at": a.created_at.isoformat(),
        })

    return report


@router.get("/transporter-performance")
async def transporter_performance(
    current_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    transporters = db.query(Transporter).all()

    report = []
    for t in transporters:
        total_bids = db.query(func.count(Bid.id)).filter(
            Bid.transporter_id == t.id, Bid.is_latest == True
        ).scalar() or 0

        wins = db.query(func.count(AuctionResult.id)).filter(
            AuctionResult.winner_id == t.id
        ).scalar() or 0

        avg_bid = db.query(func.avg(Bid.amount)).filter(
            Bid.transporter_id == t.id, Bid.is_latest == True
        ).scalar()

        win_rate = (wins / total_bids * 100) if total_bids > 0 else 0

        report.append({
            "transporter_id": t.id,
            "company_name": t.company_name,
            "city": t.city,
            "total_bids": total_bids,
            "total_wins": wins,
            "win_rate": round(win_rate, 1),
            "avg_bid_amount": round(float(avg_bid), 2) if avg_bid else 0,
            "rating": t.rating,
            "is_verified": t.is_verified,
        })

    report.sort(key=lambda x: x["win_rate"], reverse=True)
    return report


@router.get("/monthly")
async def monthly_report(
    months: int = Query(6, ge=1, le=24),
    current_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    reports = []
    now = datetime.now(timezone.utc)

    for i in range(months):
        month_start = (now.replace(day=1) - timedelta(days=30 * i)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if month_start.month == 12:
            month_end = month_start.replace(year=month_start.year + 1, month=1)
        else:
            month_end = month_start.replace(month=month_start.month + 1)

        total = db.query(func.count(Auction.id)).filter(
            Auction.created_at >= month_start, Auction.created_at < month_end
        ).scalar() or 0

        bids = db.query(func.count(Bid.id)).filter(
            Bid.submitted_at >= month_start, Bid.submitted_at < month_end
        ).scalar() or 0

        awarded = db.query(func.count(AuctionResult.id)).filter(
            AuctionResult.awarded_at >= month_start, AuctionResult.awarded_at < month_end
        ).scalar() or 0

        total_value = db.query(func.sum(AuctionResult.winning_amount)).filter(
            AuctionResult.awarded_at >= month_start, AuctionResult.awarded_at < month_end
        ).scalar() or 0

        reports.append(MonthlyReport(
            month=month_start.strftime("%Y-%m"),
            total_auctions=total,
            total_bids=bids,
            awarded_auctions=awarded,
            total_value=float(total_value),
            avg_savings_percent=0,
        ))

    reports.reverse()
    return reports


@router.get("/savings")
async def savings_report(
    current_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    results = db.query(AuctionResult).all()

    total_reserve = 0
    total_awarded = 0
    details = []

    for r in results:
        auction = db.query(Auction).filter(Auction.id == r.auction_id).first()
        if auction and auction.reserve_price:
            reserve = float(auction.reserve_price)
            awarded = float(r.winning_amount)
            savings = reserve - awarded
            savings_pct = (savings / reserve * 100) if reserve > 0 else 0

            total_reserve += reserve
            total_awarded += awarded

            winner = db.query(Transporter).filter(Transporter.id == r.winner_id).first()
            details.append({
                "auction_number": auction.auction_number,
                "route": f"{auction.pickup_location} → {auction.destination}",
                "reserve_price": reserve,
                "winning_amount": awarded,
                "savings": round(savings, 2),
                "savings_percent": round(savings_pct, 1),
                "winner": winner.company_name if winner else None,
            })

    return {
        "total_reserve": round(total_reserve, 2),
        "total_awarded": round(total_awarded, 2),
        "total_savings": round(total_reserve - total_awarded, 2),
        "avg_savings_percent": round(((total_reserve - total_awarded) / total_reserve * 100) if total_reserve > 0 else 0, 1),
        "details": details,
    }


@router.get("/export/excel/{report_type}")
async def export_excel(
    report_type: str,
    current_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(500, "openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = report_type.replace("-", " ").title()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    if report_type == "auction-summary":
        headers = ["Auction #", "Pickup", "Destination", "Material", "Vehicle", "Status",
                    "Bids", "Reserve Price", "Lowest Bid", "Winner", "Savings", "Loading Date"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        auctions = db.query(Auction).order_by(Auction.created_at.desc()).all()
        for row, a in enumerate(auctions, 2):
            result = db.query(AuctionResult).filter(AuctionResult.auction_id == a.id).first()
            lowest = db.query(func.min(Bid.amount)).filter(Bid.auction_id == a.id, Bid.is_latest == True).scalar()
            winner = None
            savings = 0
            if result:
                w = db.query(Transporter).filter(Transporter.id == result.winner_id).first()
                winner = w.company_name if w else None
                if a.reserve_price:
                    savings = float(a.reserve_price) - float(result.winning_amount)

            values = [
                a.auction_number, a.pickup_location, a.destination, a.material_type,
                a.vehicle_type, a.status, a.total_bids,
                float(a.reserve_price) if a.reserve_price else "",
                float(lowest) if lowest else "", winner or "", round(savings, 2),
                a.loading_date.strftime("%d/%m/%Y") if a.loading_date else "",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border

    elif report_type == "transporter-performance":
        headers = ["Company", "City", "Total Bids", "Wins", "Win Rate %", "Avg Bid", "Rating", "Verified"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        transporters = db.query(Transporter).all()
        for row, t in enumerate(transporters, 2):
            bids = db.query(func.count(Bid.id)).filter(Bid.transporter_id == t.id, Bid.is_latest == True).scalar() or 0
            wins = db.query(func.count(AuctionResult.id)).filter(AuctionResult.winner_id == t.id).scalar() or 0
            avg_bid = db.query(func.avg(Bid.amount)).filter(Bid.transporter_id == t.id, Bid.is_latest == True).scalar()
            win_rate = (wins / bids * 100) if bids > 0 else 0

            values = [t.company_name, t.city, bids, wins, round(win_rate, 1),
                      round(float(avg_bid), 2) if avg_bid else 0, t.rating, "Yes" if t.is_verified else "No"]
            for col, val in enumerate(values, 1):
                ws.cell(row=row, column=col, value=val)

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=TRAMS_{report_type}_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )


@router.get("/export/pdf/{report_type}")
async def export_pdf(
    report_type: str,
    current_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(500, "reportlab not installed")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    title = Paragraph(f"TRAMS - {report_type.replace('-', ' ').title()} Report", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 20))

    date_str = Paragraph(f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"])
    elements.append(date_str)
    elements.append(Spacer(1, 20))

    if report_type == "auction-summary":
        data = [["Auction #", "Route", "Material", "Status", "Bids", "Reserve", "Winner", "Savings"]]

        auctions = db.query(Auction).order_by(Auction.created_at.desc()).limit(100).all()
        for a in auctions:
            result = db.query(AuctionResult).filter(AuctionResult.auction_id == a.id).first()
            winner = None
            savings = 0
            if result:
                w = db.query(Transporter).filter(Transporter.id == result.winner_id).first()
                winner = w.company_name if w else None
                if a.reserve_price:
                    savings = float(a.reserve_price) - float(result.winning_amount)

            data.append([
                a.auction_number,
                f"{a.pickup_location[:15]} → {a.destination[:15]}",
                a.material_type[:20],
                a.status.upper(),
                str(a.total_bids),
                f"₹{float(a.reserve_price):,.0f}" if a.reserve_price else "-",
                (winner or "-")[:20],
                f"₹{savings:,.0f}" if savings else "-",
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4FF")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=TRAMS_{report_type}_{datetime.now().strftime('%Y%m%d')}.pdf"},
    )
